#!/usr/bin/env python3
"""Build the Jacobs et al. 2018 snapshots (Table 3 + Table 5).

Faithful capture per __HOWTO_make_a_snapshot.md: printed headers, values exactly
as printed (mean +/- SD in one cell, ranges as one cell with the en dash),
printed grouping rows kept, printed row order kept. Print typos are NOT fixed
here -- they are carried verbatim and flagged in the README.

The only documented deviation: the printed Species cell holds the common name on
line 1 and the italic binomial on line 2; that single cell is split into two
columns (Species / Species_binomial) so the file is machine-readable. Both
printed strings survive verbatim (invariant 3).
"""
import os

import openpyxl
from openpyxl.styles import Font, Alignment

# Derive the output folder from this file's own location so the script runs on any machine
# (__HOWTO_build_a_dataset_file.md, "Known wrinkles -> Paths"). It previously hardcoded an
# absolute sandbox path that no longer exists, which made the snapshots unregenerable.
OUT = os.path.dirname(os.path.abspath(__file__))
PM = "±"      # +/-
EN = "–"      # en dash, as printed in ranges
MU = "µ"      # micro sign

# ----------------------------------------------------------------- TABLE 3
# Estimated average lengths, areas, and volumes of layer V pyramidal and
# gigantopyramidal neuron somata in carnivore and primate species.
# Column blocks: PyrLen, GigLen, PyrArea, GigArea, PyrVol, GigVol
# each block = (n, Mean, Range) exactly as printed.
T3_TIER1 = [
    "", "", "Body mass", "Brain mass",
    f"Pyramidal neuron lengths ({MU}m)", "", "",
    f"Gigantopyramidal neuron lengths ({MU}m)", "", "",
    f"Pyramidal neuron areas ({MU}m2)", "", "",
    f"Gigantopyramidal neuron areas ({MU}m2)", "", "",
    f"Pyramidal neuron volumes ({MU}m3)", "", "",
    f"Gigantopyramidal neuron volumes ({MU}m3)", "", "",
]
T3_TIER2 = [
    "Species", "Species_binomial", "(kg)", "(g)",
    "n", "Mean", "Range", "n", "Mean", "Range",
    "n", "Mean", "Range", "n", "Mean", "Range",
    "n", "Mean", "Range", "n", "Mean", "Range",
]

T3 = [
    ["GROUP", "Carnivores"],
    ["African lion", "Panthera leo", "155.8", "223.5",
     "137", f"8.9{PM}1.5", f"5.7{EN}14.8", "15", f"20.5{PM}4.6", f"15.0{EN}29.8",
     "134", f"268.9{PM}86.1", f"113.3{EN}596.0", "16", f"1,416.2{PM}644.4", f"728.6{EN}2,858.5",
     "132", f"3,793.6{PM}1,872.5", f"1,014.5{EN}13,002.4", "16", f"46,173.6{PM}30,550.2", f"15,350.5{EN}119,782.3"],
    ["African wild dog", "Lycaon pictus", "26.3", "141.5",
     "178", f"9.1{PM}1.0", f"6.3{EN}10.8", "83", f"12.9{PM}2.0", f"10.8{EN}21.7",
     "167", f"286.4{PM}51.8", f"20.11{EN}389.4", "76", f"586.4{PM}206.7", f"391.0{EN}1,622.1",
     "207", f"4,391.3{PM}1,700.0", f"2,025.6{EN}8,996.72", "41", f"16,074.3{PM}8,689.9", f"9,109.7{EN}55,734,1"],
    ["Amur leopard", "Panthera pardus orientalis", "52.4", "125.5",
     "292", f"8.3{PM}1.1", f"6.2{EN}11.7", "29", f"19.2{PM}6.1", f"12.3{EN}29.4",
     "295", f"225.9{PM}64.1", f"122.4{EN}458.2", "29", f"1,316{PM}818.2", f"479.5{EN}2,765.8",
     "298", f"2,775.7{PM}1,291.9", f"1,051.0{EN}8,841.8", "27", f"45,239.7{PM}38,077.4", f"9,888.8{EN}117,188.8"],
    ["Asian small-clawed otter", "Amblonyx cinereus", "3.5", "38.1",
     "216", f"8.6{PM}1.6", f"6.2{EN}12.8", "31", f"16.0{PM}2.5", f"13.1{EN}23.0",
     "217", f"252.1{PM}102.2", f"124.5{EN}552.5", "29", f"872.8{PM}274.0", f"566.7{EN}1,667.0",
     "218", f"3,266.5{PM}2,056.3", f"1,006.1{EN}9,897.8", "30", f"20,560.2{PM}10,114.9", f"10,167.7{EN}51,610.2"],
    ["Banded mongoose", "Mungos mungo", "1.3", "10.5",
     "503", f"7.5{PM}0.8", f"6.0{EN}9.0", "60", f"9.8{PM}0.8", f"9.0{EN}12.0",
     "509", f"187.5{PM}38.1", f"120.0{EN}278.3", "46", f"330.3{PM}51.4", f"281.3{EN}495.8",
     "509", f"2,099.2{PM}691.7", f"1,003.2{EN}3,590.5", "31", f"5,297.7{PM}1,167.3", f"40,16.1{EN}8,954.5"],
    ["Caracal", "Caracal caracal", "11.6", "55.3",
     "282", f"8.8{PM}1.2", f"7.6{EN}12.3", "41", f"15.2{PM}2.5", f"12.7{EN}27.7",
     "285", f"258.3{PM}74.8", f"186.0{EN}520.0", "41", f"775.0{PM}317.0", f"520.9{EN}2,509.1",
     "284", f"3,415.1{PM}1,663.6", f"2,001.3{EN}9,691.3", "40", f"18,421.7{PM}14,309.5", f"10,129.5{EN}100,697.6"],
    ["European polecat", "Mustela putorius", "1.1", "8.3",
     "177", f"8.2{PM}1.5", f"6.0{EN}11.0", "44", f"12.4{PM}1.0", f"11.1{EN}14.8",
     "183", f"241.1{PM}89.3", f"119.3{EN}433.3", "32", f"533.6{PM}68.6", f"443.5{EN}725.2",
     "190", f"3,214.3{PM}1,819.4", f"1,004.3{EN}7,649.7", "27", f"10,231.0{PM}1,777.3", f"8,187.3{EN}15,925.5"],
    ["Harp seal", "Pagophilus groenlandicus", "132.3", "276.0",
     "240", f"9.1{PM}1.2", f"6.3{EN}11.0", "49", f"11.7{PM}0.7", f"11.0{EN}13.9",
     "237", f"271.7{PM}65.2", f"127.9{EN}387.9", "52", f"438.5{PM}54.3", f"390.5{EN}614.6",
     "237", f"3,585.2{PM}1233.7", f"1,104.6{EN}5,983.4", "52", f"7,296.3{PM}1,491.9", f"6,012.3{EN}11,984.2"],
    ["Northern fur seal", "Callorhinus ursinus", "135.9", "328.6",
     "433", f"9.6{PM}1.6", f"6.1{EN}12.5", "43", f"13.4{PM}1.0", f"12.5{EN}17.5",
     "447", f"319.1{PM}105.6", f"122.3{EN}549.2", "27", f"642.8{PM}106.1", f"556.4{EN}1,053.3",
     "443", f"4,667.7{PM}2,249.8", f"1,013.7{EN}9,971.2", "34", f"12,955.6{PM}3,638.7", f"10,033.7{EN}28,902.2"],
    ["Raccoon", "Procyon lotor", "6.4", "40.0",
     "541", f"8.8{PM}1.2", f"6.3{EN}13.0", "39", f"15.4{PM}2.3", f"13.0{EN}21.5",
     "551", f"254.1{PM}74.8", f"120.7{EN}548.4", "37", f"794.6{PM}253.8", f"562.4{EN}1,488.5",
     "550", f"3,232.9{PM}1,440.6", f"1,000.8{EN}9,666.8", "40", f"17,746.9{PM}9,299.3", f"10,075.0{EN}44,546.1"],
    ["Siberian tiger", "Panthera tigris altaica", "161.0", "279.3",
     "126", f"8.6{PM}1.2", f"6.4{EN}11.8", "12", f"17.0{PM}4.6", f"12.1{EN}27.8",
     "128", f"245.9{PM}72.7", f"126.3{EN}452.0", "12", f"1,040.0{PM}569.8", f"460.9{EN}2,517.8",
     "131", f"3,158.4{PM}14,71.9", f"1,045.7{EN}7,560.3", "11", f"32,289.6{PM}25,103.6", f"9,909.7{EN}99,838.6"],
    ["GROUP", "Primates"],
    ["Black-capped squirrel monkey", "Saimiri boliviensis", "0.9", "25.5",
     "146", f"6.5{PM}0.5", f"5.8{EN}7.5", "24", f"8.2{PM}0.8", f"7.7{EN}11",
     "126", f"140.7{PM}20.2", f"114.2{EN}189.5", "22", f"224.4{PM}46.4", f"191.8{EN}392.0",
     "109", f"1,360.5{PM}266.4", f"1,001.7{EN}1,949.0", "24", f"2,646.0{PM}935.7", f"2,004.6{EN}6,030.8"],
    ["Chacma baboon", "Papio ursinus", "31.0", "214.4",
     "27", f"6.5{PM}0.2", f"6.0{EN}6.9", "27", f"8.0{PM}0.8", f"7.2{EN}9.8",
     "35", f"143.8{PM}16.3", f"121.0{EN}178.3", "19", f"223.7{PM}35.8", f"182.0{EN}305.8",
     "38", f"1,364.1{PM}252.0", f"1,016.7{EN}1,909.4", "18", f"2,662.7{PM}593.6", f"2,014.3{EN}4,069.2"],
    ["Cotton-top tamarin", "Saguinus oedipus", "0.4", "8.9",
     "137", f"6.5{PM}0.5", f"5.8{EN}7.5", "35", f"8.9{PM}1.2", f"7.6{EN}12.0",
     "108", f"145.5{PM}18.3", f"118.3{EN}187.9", "37", f"260.9{PM}73.5", f"191.9{EN}473.5",
     "111", f"1,368.7{PM}239.2", f"1,000.1{EN}1,955.0", "41", f"3,349.5{PM}1,506.8", f"2,008.8{EN}8,339.8"],
    ["Golden lion tamarin", "Leontopithecus rosalia", "0.6", "12.8",
     "134", f"5.7{PM}0.5", f"4.8{EN}6.7", "27", f"7.5{PM}0.7", f"6.7{EN}9.5",
     "130", f"107.7{PM}19.8", f"77.0{EN}148.1", "32", f"183.1{PM}38.5", f"149.5{EN}299.5",
     "140", f"909.3{PM}272.5", f"511.5{EN}1,498.9", "29", f"2,171.0{PM}713.3", f"1,530.6{EN}4,147.9"],
    ["Hamadryas baboon", "Papio hamadryas", "26.0", "159.1",
     "104", f"6.5{PM}0.3", f"6.0{EN}7.0", "64", f"7.9{PM}1.3", f"7.0{EN}15.7",
     "130", f"143.1{PM}16.0", f"118.3{EN}175.0", "41", f"235.0{PM}101.1", f"176.1{EN}784.4",
     "143", f"1,385.2{PM}267.5", f"1,002.6{EN}1,980.9", "35", f"3,327.7{PM}2,644.4", f"2,001.8{EN}16,828.2"],
    ["Lar gibbon", "Hylobates lar", "5.5", "134.8",
     "97", f"7.8{PM}0.1", f"6.1{EN}8.6", "96", f"9.3{PM}0.04", f"8.6{EN}9.9",
     "123", f"205.8{PM}30.6", f"122.7{EN}247.9", "46", f"273.8{PM}13.6", f"251.8{EN}295.1",
     "159", f"2,558.9{PM}637.9", f"1,061.1{EN}3,640.8", "71", f"5,603.9{PM}3,121.4", f"3,657.9{EN}12,964.4"],
    ["Pygmy marmoset", "Cebuella pygmaea", "0.1", "4.1",
     "108", f"6.6{PM}0.3", f"6.2{EN}7.4", "42", f"8.3{PM}0.6", f"7.5{EN}9.6",
     "127", f"141.3{PM}16.4", f"120.0{EN}183.7", "40", f"229.4{PM}35.0", f"187.4{EN}295.0",
     "134", f"1,304.5{PM}229.8", f"1,002.2{EN}1,943.7", "42", f"2,757.2{PM}666.1", f"2,007.3{EN}4,313.6"],
    ["Red-tailed monkey", "Cercopithecus ascanius", "4.2", "58.4",
     "193", f"6.7{PM}0.3", f"6.2{EN}7.4", "111", f"8.1{PM}0.6", f"7.4{EN}10.5",
     "241", f"150.2{PM}20.2", f"120.7{EN}189.2", "76", f"224.5{PM}27.7", f"190.1{EN}347.6",
     "239", f"1,421.9{PM}287.1", f"1,000.6{EN}1,992.0", "83", f"2,567.1{PM}490.8", f"2,002.7{EN}4901.5"],
    ["Vervet monkey", "Chlorocebus pygerythrus", "5.6", "71.8",
     "57", f"7.2{PM}0.6", f"6.2{EN}8.4", "26", f"10.3{PM}1.9", f"8.4{EN}14.9",
     "62", f"167.1{PM}30.4", f"122.0{EN}228.9", "27", f"352.6{PM}139.9", f"237.6{EN}715.6",
     "71", f"1,698.7{PM}536.0", f"1,005.8{EN}2929.2", "26", f"5,636.5{PM}3,458.1", f"3,033.8{EN}15,219.6"],
]

T3_FOOT = ('a Each species is represented by one animal. The standard deviation and the range '
           'observed as well as the number of cells measured are provided. Estimates of area '
           'were calculated using the “nucleator” probe of the StereoInvestigator '
           'software (see text for details).')

# ----------------------------------------------------------------- TABLE 5
T5_HEAD = ["Neuron type", "na", "Volb", "TDLc", "MSLd", "DSCe", "DSNf", "DSDg",
           "SoSizeh", "SoDepthi"]

def r(nt, n, vol, tdl, msl, dsc, dsn, dsd, sosize, sodepth):
    return [nt, n, vol, tdl, msl, dsc, dsn, dsd, sosize, sodepth]

def pm(a, b):
    return f"{a}{PM}{b}"

T5 = [
    ["GROUP", "Artiodactyls"], ["SPECIES", "Blue wildebeest"],
    r("Superficial", "19", pm("13,411", "1,527"), pm("5,640", "414"), pm("77", "3"), pm("75", "6"), pm("2,224", "220"), pm("0.30", "0.02"), pm("425", "30"), pm("759", "68")),
    r("Deep", "13", pm("22,158", "2,526"), pm("6,086", "306"), pm("86", "4"), pm("73", "5"), pm("2,236", "232"), pm("0.30", "0.03"), pm("479", "38"), pm("1,387", "105")),
    r("Gigantopyramidal", "26", pm("51,400", "3,386"), pm("5,666", "266"), pm("101", "4"), pm("57", "3"), pm("3,244", "219"), pm("0.50", "0.02"), pm("882", "45"), pm("1,361", "47")),
    ["SPECIES", "Giraffe"],
    r("Superficial", "11", pm("16,059", "2,157"), pm("5,177", "319"), pm("81", "3"), pm("64", "3"), pm("4,955", "383"), pm("0.84", "0.06"), pm("301", "27"), pm("658", "31")),
    r("Deep", "10", pm("16,222", "1,927"), pm("5,980", "659"), pm("85", "6"), pm("70", "5"), pm("4,370", "579"), pm("0.58", "0.03"), pm("408", "31"), pm("1,218", "37")),
    r("Gigantopyramidal", "5", pm("73,427", "15,718"), pm("8,397", "903"), pm("100", "4"), pm("84", "8"), pm("7,794", "902"), pm("0.75", "0.04"), pm("1,143", "112"), pm("1,513", "92")),
    ["SPECIES", "Kudu"],
    r("Superficial", "16", pm("7,445", "633"), pm("3,655", "269"), pm("71", "3"), pm("51", "3"), pm("2,068", "238"), pm("0.40", "0.04"), pm("323", "13"), pm("948", "18")),
    r("Deep", "13", pm("10,316", "1,309"), pm("3,771", "244"), pm("72", "5"), pm("54", "4"), pm("1,726", "199"), pm("0.30", "0.03"), pm("357", "30"), pm("1,468", "91")),
    r("Gigantopyramidal", "7", pm("25,210", "2,779"), pm("5,191", "592"), pm("87", "4"), pm("61", "9"), pm("2,135", "173"), pm("0.30", "0.04"), pm("745", "71"), pm("1,724", "102")),
    ["GROUP", "Caniforms"], ["SPECIES", "African wild dog"],
    r("Superficial", "41", pm("10,853", "677"), pm("4,681", "222"), pm("79", "2"), pm("60", "3"), pm("2,343", "147"), pm("0.40", "0.02"), pm("376", "18"), pm("759", "26")),
    r("Deep", "31", pm("11,050", "639"), pm("4,707", "259"), pm("83", "3"), pm("57", "3"), pm("1,892", "193"), pm("0.30", "0.02"), pm("428", "18"), pm("1,385", "62")),
    r("Gigantopyramidal", "9", pm("27,317", "3,821"), pm("5,794", "499"), pm("79", "4"), pm("72", "5"), pm("3,038", "399"), pm("0.42", "0.04"), pm("2,662", "261"), pm("1,554", "45")),
    ["SPECIES", "Domestic dog"],
    r("Superficial", "12", pm("9,446", "1,147"), pm("3,112", "165"), pm("58", "3"), pm("55", "4"), pm("929", "95"), pm("0.30", "0.02"), pm("430", "44"), pm("718", "41")),
    r("Deep", "8", pm("12,572", "1,700"), pm("3,429", "366"), pm("68", "5"), pm("51", "5"), pm("929", "122"), pm("0.20", "0.03"), pm("629", "67"), pm("1,230", "75")),
    r("Gigantopyramidal", "17", pm("37,564", "3,807"), pm("3,524", "238"), pm("73", "4"), pm("49", "3"), pm("1,225", "149"), pm("0.30", "0.03"), pm("1,876", "187"), pm("1,402", "66")),
    ["GROUP", "Diprotodont"], ["SPECIES", "Bennett’s wallaby"],
    r("Superficial", "10", pm("4,951", "636"), pm("4,739", "381"), pm("90", "5"), pm("53", "4"), pm("2,697", "298"), pm("0.48", "0.04"), pm("258", "13"), pm("540", "37")),
    r("Deep", "13", pm("14,151", "2,023"), pm("4,705", "619"), pm("85", "6"), pm("52", "4"), pm("2,753", "467"), pm("0.38", "0.03"), pm("437", "31"), pm("1,052", "33")),
    ["GROUP", "Feliforms"], ["SPECIES", "Banded mongoose"],
    r("Superficial", "10", pm("6,400", "1,010"), pm("5,092", "539"), pm("72", "5"), pm("71", "6"), pm("3,011", "442"), pm("0.48", "0.04"), pm("279", "21"), pm("622", "29")),
    r("Deep", "14", pm("11,551", "1,404"), pm("4,931", "327"), pm("81", "5"), pm("61", "3"), pm("3,120", "251"), pm("0.51", "0.03"), pm("484", "74"), pm("992", "34")),
    ["SPECIES", "Caracal"],
    r("Superficial", "10", pm("7,962", "1,117"), pm("6,008", "464"), pm("82", "4"), pm("75", "6"), pm("3,395", "470"), pm("0.46", "0.05"), pm("370", "29"), pm("670", "54")),
    r("Deep", "9", pm("15,765", "3,052"), pm("5,353", "329"), pm("88", "5"), pm("61", "3"), pm("3,661", "306"), pm("0.53", "0.02"), pm("448", "45"), pm("1,267", "64")),
    r("Gigantopyramidal", "10", pm("67,799", "15,311"), pm("7,222", "712"), pm("88", "4"), pm("81", "6"), pm("1,187", "214"), pm("0.15", "0.05"), pm("2,000", "182"), pm("1,199", "47")),
    ["SPECIES", "Clouded leopard"],
    r("Superficial", "5", pm("8,452", "1,563"), pm("3,521", "255"), pm("62", "7"), pm("59", "6"), pm("1,353", "37"), pm("0.35", "0.02"), pm("306", "27"), pm("571", "50")),
    r("Deep", "1", "4,203", "4,436", "78", "57", "1,170", "0.22", "294", "1,435"),
    r("Gigantopyramidal", "6", pm("61,253", "11,929"), pm("1,179", "205"), pm("63", "7"), pm("18", "2"), pm("240", "38"), pm("0.20", "0.03"), pm("3,720", "352"), pm("1,363", "53")),
    ["SPECIES", "Siberian tiger"],
    r("Superficial", "6", pm("20,182", "5,677"), pm("6,215", "247"), pm("83", "8"), pm("77", "5"), pm("2,257", "308"), pm("0.28", "0.04"), pm("439", "85"), pm("752", "71")),
    r("Deep", "3", pm("16,066", "1,812"), pm("5,589", "352"), pm("97", "14"), pm("60", "10"), pm("1,496", "98"), pm("0.19", "0.02"), pm("350", "62"), pm("1,486", "244")),
    r("Gigantopyramidal", "8", pm("94,701", "27,983"), pm("5,513", "1,003"), pm("128", "7"), pm("42", "6"), pm("1,005", "157"), pm("0.16", "0.02"), pm("2,844", "366"), pm("1,546", "29")),
    ["SPECIES", "African lion"],
    r("Superficial", "10", pm("9,762", "1,747"), pm("4,998", "259"), pm("77", "3"), pm("65", "3"), pm("1,415", "159"), pm("0.23", "0.03"), pm("358", "40"), pm("799", "45")),
    r("Deep", "10", pm("17,673", "4,006"), pm("5,800", "555"), pm("83", "4"), pm("69", "5"), pm("1,908", "306"), pm("0.24", "0.02"), pm("525", "87"), pm("1,376", "73")),
    r("Gigantopyramidal", "10", pm("88,381", "8,281"), pm("8,600", "814"), pm("120", "8"), pm("72", "6"), pm("1,099", "154"), pm("0.10", "0.01"), pm("2,824", "284"), pm("1,558", "40")),
    ["GROUP", "Lagomorph"], ["SPECIES", "Flemish giant rabbit"],
    r("Superficial", "19", pm("4,261", "299"), pm("2,861", "207"), pm("67", "2"), pm("43", "3"), pm("1,365", "113"), pm("0.40", "0.02"), pm("248", "12"), pm("762", "27")),
    r("Deep", "18", pm("6,353", "797"), pm("3,522", "313"), pm("66", "3"), pm("52", "4"), pm("1,115", "125"), pm("0.20", "0.02"), pm("350", "19"), pm("1,322", "34")),
    ["GROUP", "Perissodactyls"], ["SPECIES", "Mountain zebra"],
    r("Superficial", "5", pm("12,487", "1,144"), pm("5,352", "529"), pm("92", "4"), pm("58", "3"), pm("2,221", "701"), pm("0.30", "0.08"), pm("423", "31"), pm("796", "57")),
    r("Deep", "6", pm("17,925", "3,584"), pm("5,433", "524"), pm("93", "5"), pm("60", "7"), pm("2,109", "375"), pm("0.30", "0.03"), pm("464", "46"), pm("1,427", "131")),
    r("Gigantopyramidal", "21", pm("60,377", "3,604"), pm("4,871", "340"), pm("89", "2"), pm("55", "4"), pm("2,280", "202"), pm("0.40", "0.03"), pm("1,260", "70"), pm("1,434", "60")),
    ["SPECIES", "Plains zebra"],
    r("Superficial", "10", pm("13,088", "1,963"), pm("5,105", "469"), pm("79", "1"), pm("65", "6"), pm("2,896", "497"), pm("0.45", "0.04"), pm("377", "42"), pm("818", "50")),
    r("Deep", "13", pm("14,339", "1,613"), pm("5,015", "407"), pm("87", "4"), pm("58", "5"), pm("2,637", "307"), pm("0.41", "0.02"), pm("455", "28"), pm("1,168", "59")),
    r("Gigantopyramidal", "15", pm("51,879", "5,012"), pm("3,970", "395"), pm("104", "4"), pm("39", "4"), pm("1,976", "271"), pm("0.37", "0.03"), pm("870", "79"), pm("1,188", "21")),
    ["GROUP", "Primates"], ["SPECIES", "Ring-tailed lemur"],
    r("Superficial", "9", pm("5,729", "672"), pm("4,571", "365"), pm("65", "3"), pm("70", "5"), pm("2,652", "329"), pm("0.50", "0.03"), pm("225", "14"), pm("526", "33")),
    r("Deep", "1", "2,611", "3,118", "82", "38", "1,235", "0.28", "172", "877"),
    r("Gigantopyramidal", "2", pm("22,159", "5,598"), pm("4,787", "691"), pm("69", "0.04"), pm("69", "10"), pm("2,308", "200"), pm("0.46", "0.03"), pm("1,007", "37"), pm("1,188", "39")),
    ["SPECIES", "Golden lion tamarin"],
    r("Superficial", "10", pm("2,650", "512"), pm("4,476", "330"), pm("64", "2"), pm("71", "6"), pm("1,986", "160"), pm("0.42", "0.03"), pm("185", "21"), pm("715", "52")),
    r("Deep", "10", pm("5,111", "856"), pm("5,029", "447"), pm("70", "4"), pm("72", "6"), pm("2,211", "257"), pm("0.38", "0.03"), pm("268", "34"), pm("1,255", "76")),
    r("Gigantopyramidal", "19", pm("26,474", "2,508"), pm("6,286", "399"), pm("89", "3"), pm("73", "6"), pm("2,008", "123"), pm("0.28", "0.02"), pm("954", "60"), pm("1,230", "38")),
    ["SPECIES", "Chacma baboon"],
    r("Superficial", "10", pm("8,294", "1,142"), pm("6,511", "494"), pm("68", "3"), pm("95", "5"), pm("4,468", "546"), pm("0.70", "0.03"), pm("282", "21"), pm("615", "54")),
    r("Deep", "10", pm("10,491", "2,355"), pm("5,968", "615"), pm("77", "4"), pm("77", "6"), pm("3,347", "387"), pm("0.50", "0.03"), pm("389", "59"), pm("1,505", "114")),
    r("Gigantopyramidal", "9", pm("45,638", "10,333"), pm("9,524", "621"), pm("90", "4"), pm("107", "7"), pm("4,002", "837"), pm("0.40", "0.05"), pm("1,018", "167"), pm("1,763", "157")),
    ["SPECIES", "Human"],
    r("Superficial", "10", pm("12,717", "1,458"), pm("6,121", "487"), pm("66", "3"), pm("92", "4"), pm("2,213", "227"), pm("0.32", "0.02"), pm("452", "46"), pm("1,255", "97")),
    r("Deep", "10", pm("11,300", "1,949"), pm("5,329", "516"), pm("80", "3"), pm("68", "7"), pm("2,108", "313"), pm("0.33", "0.02"), pm("454", "60"), pm("2,146", "96")),
    r("Gigantopyramidal", "17", pm("29,412", "2,804"), pm("7,706", "503"), pm("91", "4"), pm("87", "7"), pm("1,736", "275"), pm("0.20", "0.03"), pm("969", "52"), pm("1,876", "68")),
    ["GROUP", "Rodent"], ["SPECIES", "Rat"],
    r("Superficial", "10", pm("4,126", "346"), pm("3,785", "257"), pm("61", "2"), pm("62", "4"), pm("2,252", "184"), pm("0.50", "0.02"), pm("173", "11"), pm("468", "22")),
    r("Deep", "10", pm("10,015", "1,153"), pm("4,605", "262"), pm("71", "3"), pm("65", "4"), pm("2,174", "198"), pm("0.41", "0.03"), pm("362", "22"), pm("1,208", "47")),
]

T5_FOOT = [
    "a Number of cells traced.",
    f"b Volume in {MU}m3.",
    f"c Length in {MU}m.",
    f"d Average length of dendritic segments in {MU}m.",
    "e Number of segments per neuron.",
    "f Number of spines per neuron.",
    f"g Number of spines per {MU}m of dendritic length.",
    f"h Soma size in {MU}m2.",
    f"i Soma depth in {MU}m from the pial surface.",
]


def write_t3():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table3"
    ws.append([("TABLE 3  Estimated average lengths, areas, and volumes of layer V pyramidal and "
                "gigantopyramidal neuron somata in carnivore and primate species (in alphabetical "
                "order by species)a")])
    ws.append(T3_TIER1); ws.append(T3_TIER2)
    for row in T3:
        if row[0] == "GROUP":
            ws.append([row[1]])
        else:
            ws.append(row)
    ws.append([]); ws.append([T3_FOOT])
    for c in ws[2] + ws[3]:
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 26
    for col in "CDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[col].width = 17
    wb.save(f"{OUT}/Jacobs_etal_2018_Table3_snapshot.xlsx")


def write_t5():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Table5"
    ws.append([("TABLE 5  Summary statistics for each neuron type across sampled species in "
                "morphological analysis (in alphabetical order by taxonomic group)")])
    ws.append(T5_HEAD)
    for row in T5:
        if row[0] == "GROUP":
            ws.append([row[1]])
        elif row[0] == "SPECIES":
            ws.append(["", row[1]])   # printed as an indented species sub-heading
        else:
            ws.append(row)
    ws.append([])
    for f in T5_FOOT:
        ws.append([f])
    for c in ws[2]:
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 22
    for col in "CDEFGHIJ":
        ws.column_dimensions[col].width = 18
    wb.save(f"{OUT}/Jacobs_etal_2018_Table5_snapshot.xlsx")


if __name__ == "__main__":
    write_t3(); write_t5()
    n3 = len([r for r in T3 if r[0] != "GROUP"])
    n5sp = len([r for r in T5 if r[0] == "SPECIES"])
    n5rows = len([r for r in T5 if r[0] not in ("GROUP", "SPECIES")])
    print(f"Table3: {n3} species rows")
    print(f"Table5: {n5sp} species, {n5rows} neuron-type rows")
