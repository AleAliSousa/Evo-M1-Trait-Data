#!/usr/bin/env python3
"""Build the Nudo et al. 1995 snapshots (TABLE 1 - TABLE 5).

Source: Nudo, R. J., Sutherland, D. P., & Masterton, R. B. (1995). Variation and
evolution of mammalian corticospinal somata with special reference to primates.
J Comp Neurol 358(2):181-205.  DOI 10.1002/cne.903580203.
PDF in this folder: "Nudo-1995-Variation and evolution of mammalian.pdf".

Printed / scanned source -> a snapshot is REQUIRED (__HOWTO_build_a_dataset_file.md
invariant 1). The PDF carries an OCR text layer, and that text layer is NOT reliable
(it renders "Rattus norvegicus" as "Raltus noruegicus", "Macaca" as "Maraca",
"cells/mm2" as "cells/mm'", "0.3" as "03", "94.0" as "94.1)"). Every cell below was
therefore read off 300-dpi renders of the printed pages, not off the text layer:

    TABLE 1  p. 183 (PDF p. 3), left column
    TABLE 2  p. 185 (PDF p. 5), full width
    TABLE 3  p. 187 (PDF p. 7), right column, above Table 4
    TABLE 4  p. 187 (PDF p. 7), right column
    TABLE 5  p. 188 (PDF p. 8), full width

Faithful capture per __HOWTO_make_a_snapshot.md: printed caption, printed header
tiers, printed row order, printed footnotes, printed thousands separators, printed
missing-value token "n/a", printed zeros, and printed typos - all verbatim. Nothing
is corrected here; corrections happen in the .R with a comment (golden rule).

Print typos deliberately carried verbatim (see the READMEs):
  * TABLE 2, Cat row      -> order printed "Camivora"  (all other tables: "Carnivora")
  * TABLE 2/4/5, Rabbit   -> order printed "Lagamorpha" (TABLE 3 prints "Lagomorpha")
  * TABLE 2, Hedgehog/H.a -> ipsilateral total printed in SQUARE brackets "615 [1,415]"
                             although footnote 2 says "parentheses"
  * TABLE 5, header       -> "Thickness (um)" but the values are millimetres (the text,
                             p. 187, calls the same numbers "slow loris (0.71 mm)")

Documented deviations (the only two):
  1. The printed headers are 2-4 stacked physical lines. They are flattened into two
     header rows (tier 1 = label, tier 2 = unit / sub-label) so the sheet is machine
     readable. No printed word is dropped.
  2. TABLE 1 common names wrap across up to three printed lines
     ("Nine-banded / armadillo", "Common mar- / moset"). They are joined with a single
     space, and a line-break hyphen ("mar-") is closed up. The scientific names, which
     never wrap, are verbatim - including the printed genus abbreviations
     "E. europaeus" and "M. mulatta".
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Font

OUT = os.path.dirname(os.path.abspath(__file__))
MU = "µ"          # micro sign, as printed in "(um)"
SUP2 = "²"        # superscript two, as printed in "(cells/mm2)"
SUP3 = "³"        # superscript three, as printed in "(cells/mm3)"

# --------------------------------------------------------------- TABLE 1
T1_CAPTION = ("TABLE 1. Body Weight, Brain Weight, and Neocortical Surface Area "
              "for Species and Genera Represented")
T1_TIER1 = ["Common name", "Scientific name", "Body weight", "Brain weight",
            "Cortical surface area"]
T1_TIER2 = ["", "", "(g)", "(mg)", "(mm" + SUP2 + ")"]
T1 = [
    ["Albino rat",                 "Rattus norvegicus",         "268",   "1,500",  "152"],
    ["Nine-banded armadillo",      "Dasypus novemcinctus",      "5,000", "7,540",  "308"],
    ["Lesser bushbaby",            "Galago senegalensis",       "160",   "2,400",  "436"],
    ["Domestic cat",               "Felis catus",               "3,068", "19,080", "2,301"],
    ["Eastern gray squirrel",      "Sciurus carolinensis",      "400",   "4,930",  "388"],
    ["13-lined ground squirrel",   "Citellus tridecemlineatus", "150",   "1,530",  "188"],
    ["West African hedgehog",      "Erinaceus albiventris",     "410",   "1,300",  "81"],
    ["European hedgehog",          "E. europaeus",              "315",   "2,260",  "193"],
    ["Desert hedgehog",            "Hemiechinus auritus",       "203",   "965",    "119"],
    ["Rock hyrax",                 "Procavia capensis",         "2,500", "12,420", "836"],
    ["Least shrew",                "Cryptotis parva",           "8",     "80",     "18"],
    ["Common marmoset",            "Callithrix jacchus",        "287",   "5,480",  "892"],
    ["Eastern American mole",      "Scalopus aquaticus",        "53",    "570",    "87"],
    ["Green monkey",               "Cercopithecus aethiops",    "1,370", "36,540", "3,616"],
    ["Crab-eating macaque",        "Macaca fascicularis",       "1,932", "38,030", "5,548"],
    ["Rhesus macaque",             "M. mulatta",                "3,300", "53,700", "7,546"],
    ["Squirrel monkey",            "Saimiri sciureus",          "691",   "17,040", "2,150"],
    ["North American opossum",     "Didelphis virginiana",      "2,727", "4,360",  "316"],
    ["Gray, short-tailed opossum", "Monodelphis domesticus",    "83",    "790",    "68"],
    ["Rabbit",                     "Oryctolagus cuniculus",     "5,230", "7,570",  "634"],
    ["Raccoon",                    "Procyon lotor",             "4,091", "31,870", "3,669"],
    ["Slow loris",                 "Nycticebus coucang",        "965",   "7,880",  "837"],
    ["Tree shrew",                 "Tupaia glis",               "230",   "2,480",  "290"],
    ["Pine vole",                  "Pitymys pinetorum",         "16",    "370",    "55"],
]
T1_FOOT = []

# --------------------------------------------------------------- TABLE 2
T2_CAPTION = "TABLE 2. Number and Percentage of Corticospinal Somata in Each Region"
T2_TIER1 = ["", "Maximum", "Soma", "Correction", "Corrected",
            "Region", "", "", "", "", "", "", "", "", ""]
T2_TIER2 = ["Animal/G.s." + "¹" + "/Order", "profile #" + SUP2,
            "diameter (" + MU + "m)", "term", "#CSN",
            "#A", "%A", "#B", "%B", "#C", "%C", "#C'", "%C'", "# Other", "% Other"]
T2 = [
    ["Albino rat/R.n./Rodentia",      "21,825",         "19.52", "0.837", "18,268",
     "15,363", "84.1", "749",   "4.1",  "0",     "0.0", "2,119", "11.6", "37",  "0.2"],
    ["Armadillo/D.n./Edentata",       "12,505",         "19.57", "0.836", "10,454",
     "10,088", "96.5", "366",   "3.5",  "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Bushbaby/G.s./Primates",        "45,335",         "18.33", "0.845", "38,308",
     "34,516", "90.1", "3,294", "8.6",  "421",   "1.1", "0",     "0.0",  "77",  "0.2"],
    ["Cat/F.c./Camivora",             "41,820",         "26.61", "0.790", "33,038",
     "29,866", "90.4", "2,412", "7.3",  "0",     "0.0", "0",     "0.0",  "760", "2.3"],
    ["Gray squirrel/S.c./Rodentia",   "19,695",         "21.89", "0.820", "16,150",
     "13,308", "82.4", "1,308", "8.1",  "0",     "0.0", "1,486", "9.2",  "48",  "0.3"],
    ["Ground squirrel/C.t./Rodentia", "24,775",         "18.30", "0.845", "20,935",
     "18,695", "89.3", "1,151", "5.5",  "0",     "0.0", "1,089", "5.2",  "0",   "0.0"],
    ["Hedgehog/E.a./Insectivora",     "765 (2,135)",    "12.57", "0.888", "1,896",
     "1,644",  "86.7", "252",   "13.3", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Hedgehog/E.e./Insectivora",     "4,255 (5,545)",  "19.42", "0.837", "4,641",
     "3,550",  "76.5", "1,091", "23.5", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Hedgehog/H.a./Insectivora",     "615 [1,415]",    "16.76", "0.856", "1,211",
     "957",    "79.0", "254",   "21.0", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Hyrax/P.c./Hyracoidea",         "5,990 (25,895)", "25.42", "0.797", "20,638",
     "19,008", "92.1", "1,548", "7.5",  "0",     "0.0", "0",     "0.0",  "82",  "0.4"],
    ["Least shrew/C.p./Insectivora",  "474",            "12.08", "0.892", "423",
     "348",    "82.3", "75",    "17.7", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Marmoset/C.j./Primates",        "31,540",         "21.06", "0.826", "26,052",
     "23,890", "91.7", "1,693", "6.5",  "469",   "1.8", "0",     "0.0",  "0",   "0.0"],
    ["Mole/S.a./Insectivora",         "5,900 (14,210)", "11.01", "0.901", "12,803",
     "9,730",  "76.0", "3,073", "24.0", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Monkey/C.a./Primates",          "75,310",         "27.14", "0.787", "59,269",
     "55,179", "93.1", "2,252", "3.8",  "1,600", "2.7", "0",     "0.0",  "237", "0.4"],
    ["Monkey/M.f./Primates",          "46,260",         "31.01", "0.763", "35,296",
     "32,331", "91.6", "1,588", "4.5",  "1,377", "3.9", "0",     "0.0",  "0",   "0.0"],
    ["Monkey/M.m./Primates",          "113,230",        "34.50", "0.743", "84,130",
     "82,447", "98.0", "1,094", "1.3",  "589",   "0.7", "0",     "0.0",  "0",   "0.0"],
    ["Monkey/S.s./Primates",          "50,080",         "23.40", "0.810", "40,565",
     "36,792", "90.7", "2,272", "5.6",  "1,420", "3.5", "0",     "0.0",  "81",  "0.2"],
    ["Opossum/D.v./Marsupialia",      "10,095",         "17.03", "0.854", "8,621",
     "8,207",  "95.2", "379",   "4.4",  "0",     "0.0", "0",     "0.0",  "35",  "0.4"],
    ["Opossum/M.d./Marsupialia",      "550",            "12.31", "0.890", "490",
     "397",    "81.1", "93",    "18.9", "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Rabbit/O.c./Lagamorpha",        "6,780",          "22.70", "0.815", "5,526",
     "4,227",  "76.5", "663",   "12.0", "0",     "0.0", "636",   "11.5", "0",   "0.0"],
    ["Raccoon/P.l./Carnivora",        "127,670",        "22.22", "0.818", "104,434",
     "101,928", "97.6", "2,402", "2.3", "0",     "0.0", "0",     "0.0",  "104", "0.1"],
    ["Slow loris/N.c./Primates",      "26,360",         "23.14", "0.812", "21,404",
     "20,013", "93.5", "1,049", "4.9",  "342",   "1.6", "0",     "0.0",  "0",   "0.0"],
    ["Tree shrew/T.g./Scandentia",    "8,394",          "20.80", "0.828", "6,950",
     "6,672",  "96.0", "278",   "4.0",  "0",     "0.0", "0",     "0.0",  "0",   "0.0"],
    ["Vole/P.p./Rodentia",            "7,120",          "14.74", "0.872", "6,209",
     "4,483",  "72.2", "459",   "7.4",  "0",     "0.0", "1,267", "20.4", "0",   "0.0"],
]
T2_FOOT = [
    "¹G.s., Genus specie; see Table 1.",
    (SUP2 + "Numbers in parentheses represent the total number of CS somata in "
     "ipsilateral hemisphere in species in which the ipsilateral total exceeded the "
     "contralateral total. Numbers to the left represent the total number of CS "
     "somata in the contralateral hemisphere."),
]

# --------------------------------------------------------------- TABLE 3
T3_CAPTION = "TABLE 3. Number of Corticospinal Somata by Taxonomic Order"
T3_TIER1 = ["", "#", "Mean", "Mean", "Mean", "Mean", "Mean", "Mean", "Mean"]
T3_TIER2 = ["Order", "Species", "# CSN" + "¹", "#A", "#B", "#C", "#C'", "%A", "%B"]
T3 = [
    ["Carnivora",   "2", "68,736", "65,897", "2,407", "0",   "0",     "94.0", "4.8"],
    ["Edentata",    "1", "10,454", "10,088", "366",   "0",   "0",     "96.5", "3.5"],
    ["Hyracoidea",  "1", "20,638", "19,008", "1,548", "0",   "0",     "92.1", "7.5"],
    ["Insectivora", "5", "4,195",  "3,246",  "949",   "0",   "0",     "80.1", "19.9"],
    ["Lagomorpha",  "1", "5,526",  "4,227",  "663",   "0",   "636",   "76.5", "12.0"],
    ["Marsupialia", "2", "4,556",  "4,302",  "236",   "0",   "0",     "88.2", "11.7"],
    ["Primates",    "7", "43,575", "40,738", "1,892", "888", "0",     "92.7", "5.0"],
    ["Rodentia",    "4", "15,391", "12,962", "917",   "0",   "1,490", "82.0", "6.3"],
    ["Scandentia",  "1", "6,950",  "6,672",  "278",   "0",   "0",     "96.0", "4.0"],
]
T3_FOOT = ["¹Numbers of CS somata are stereologically corrected values."]

# --------------------------------------------------------------- TABLE 4
# Caption unit as printed is a real superscript two ("cells/mm2"); the PDF's OCR text
# layer mangles it to "cells/mm'". Verified on a 300-dpi render of p. 187.
T4_CAPTION = ("TABLE 4. Maximum Surface Density of Corticospinal Somata "
              "(cells/mm" + SUP2 + ")")
T4_TIER1 = ["", "Region", "", "", ""]
T4_TIER2 = ["Animal/G.s./Order", "A", "B", "C", "C'"]
T4 = [
    ["Albino rat/R.n./Rodentia",      "2,250", "592",   "0",     "1,606"],
    ["Armadillo/D.n./Edentata",       "736",   "201",   "0",     "0"],
    ["Bushbaby/G.s./Primates",        "4,030", "1,323", "1,504", "0"],
    ["Cat/F.c./Carnivora",            "1,677", "189",   "0",     "0"],
    ["Gray squirrel/S.c./Rodentia",   "1,664", "574",   "0",     "976"],
    ["Ground squirrel/C.t./Rodentia", "3,310", "602",   "0",     "1,143"],
    ["Hedgehog/E.a./Insectivora",     "848",   "326",   "0",     "0"],
    ["Hedgehog/E.e./Insectivora",     "771",   "356",   "0",     "0"],
    ["Hedgehog/H.a./Insectivora",     "676",   "184",   "0",     "0"],
    ["Hyrax/P.c./Hyracoidea",         "1,148", "446",   "0",     "0"],
    ["Least shrew/C.p./Insectivora",  "706",   "107",   "0",     "0"],
    ["Marmoset/C.j./Primates",        "1,305", "435",   "638",   "0"],
    ["Mole/S.a./Insectivora",         "3,472", "1,335", "0",     "0"],
    ["Monkey/C.a./Primates",          "900",   "228",   "464",   "0"],
    ["Monkey/M.f./Primates",          "300",   "128",   "128",   "0"],
    ["Monkey/M.m./Primates",          "408",   "99",    "161",   "0"],
    ["Monkey/S.s./Primates",          "760",   "253",   "535",   "0"],
    ["Opossum/D.v./Marsupialia",      "857",   "429",   "0",     "0"],
    ["Opossum/M.d./Marsupialia",      "655",   "393",   "0",     "0"],
    ["Rabbit/O.c./Lagamorpha",        "540",   "256",   "0",     "170"],
    ["Raccoon/P.l./Carnivora",        "786",   "171",   "0",     "0"],
    ["Slow loris/N.c./Primates",      "875",   "282",   "226",   "0"],
    ["Tree shrew/T.g./Scandentia",    "1,790", "429",   "0",     "0"],
    ["Vole/P.p./Rodentia",            "2,339", "759",   "0",     "1,075"],
]
T4_FOOT = []

# --------------------------------------------------------------- TABLE 5
T5_CAPTION = ("TABLE 5. Nine Other Morphological Characteristics of "
              "Corticospinal Somata")
T5_TIER1 = ["", "Avg surface density", "Thickness", "Maximum volume density",
            "Column height", "Concentration", "Concentration (large cells; %)",
            "Rostral/caudal", "Condensation", "% Medial-lateral distribution"]
T5_TIER2 = ["Animal/G.s./Order", "(cells/mm" + SUP2 + ")", "(" + MU + "m)",
            "(cells/mm" + SUP3 + ")", "(cells)", "(%)", "", "", "", ""]
T5 = [
    ["Albino rat/R.n./Rodentia",      "639", "0.38", "6,475",  "5.02", "26.7", "78.7", "0.89",  "93.0", "29"],
    ["Armadillo/D.n./Edentata",       "140", "0.18", "4,089",  "2.51", "35.8", "n/a",  "0.80",  "84.7", "29"],
    ["Bushbaby/G.s./Primates",        "830", "0.38", "11,918", "7.61", "34.4", "91.6", "3.23",  "96.9", "27"],
    ["Cat/F.c./Carnivora",            "173", "0.17", "11,526", "2.37", "36.7", "69.0", "6.67",  "99.4", "65"],
    ["Gray squirrel/S.c./Rodentia",   "414", "0.36", "5,283",  "3.28", "28.3", "73.5", "2.86",  "95.0", "30"],
    ["Ground squirrel/C.t./Rodentia", "800", "0.38", "9,751",  "6.76", "36.8", "89.8", "1.10",  "93.7", "33"],
    ["Hedgehog/E.a./Insectivora",     "257", "0.17", "5,432",  "2.66", "18.8", "42.9", "0.85",  "94.1", "56"],
    ["Hedgehog/E.e./Insectivora",     "195", "0.35", "2,640",  "5.02", "28.4", "55.2", "0.93",  "92.7", "32"],
    ["Hedgehog/H.a./Insectivora",     "138", "0.19", "3,965",  "2.57", "15.7", "61.0", "1.22",  "96.8", "38"],
    ["Hyrax/P.c./Hyracoidea",         "258", "0.35", "3,280",  "4.00", "n/a",  "n/a",  "1.39",  "84.5", "64"],
    ["Least shrew/C.p./Insectivora",  "245", "0.15", "5,116",  "2.68", "10.7", "n/a",  "1.00",  "93.1", "50"],
    ["Marmoset/C.j./Primates",        "355", "0.35", "4,248",  "4.13", "21.8", "80.5", "1.82",  "98.5", "33"],
    ["Mole/S.a./Insectivora",         "802", "0.21", "17,848", "7.21", "34.1", "52.6", "1.41",  "92.2", "35"],
    ["Monkey/C.a./Primates",          "106", "0.37", "2,849",  "2.36", "19.6", "63.3", "5.26",  "99.0", "78"],
    ["Monkey/M.f./Primates",          "76",  "0.26", "1,372",  "2.29", "9.1",  "79.5", "10.00", "98.9", "54"],
    ["Monkey/M.m./Primates",          "108", "0.41", "1,196",  "2.23", "13.2", "86.1", "1.64",  "94.3", "78"],
    ["Monkey/S.s./Primates",          "178", "0.38", "2,302",  "2.43", "13.9", "69.7", "2.70",  "94.6", "68"],
    ["Opossum/D.v./Marsupialia",      "298", "0.20", "4,781",  "2.56", "23.1", "65.2", "1.16",  "83.7", "55"],
    ["Opossum/M.d./Marsupialia",      "168", "0.10", "7,125",  "2.67", "15.9", "43.4", "1.25",  "78.7", "54"],
    ["Rabbit/O.c./Lagamorpha",        "116", "0.46", "1,348",  "2.45", "20.8", "71.4", "2.13",  "94.8", "39"],
    ["Raccoon/P.l./Carnivora",        "103", "0.38", "2,368",  "3.27", "27.9", "87.5", "3.23",  "98.1", "61"],
    ["Slow loris/N.c./Primates",      "263", "0.71", "1,417",  "4.06", "12.6", "79.1", "3.45",  "94.0", "56"],
    ["Tree shrew/T.g./Scandentia",    "474", "0.31", "6,555",  "4.97", "28.7", "76.9", "2.86",  "90.9", "64"],
    ["Vole/P.p./Rodentia",            "685", "0.33", "7,821",  "6.10", "26.3", "81.8", "1.18",  "91.0", "40"],
]
T5_FOOT = []

TABLES = {
    "TABLE1": (T1_CAPTION, T1_TIER1, T1_TIER2, T1, T1_FOOT),
    "TABLE2": (T2_CAPTION, T2_TIER1, T2_TIER2, T2, T2_FOOT),
    "TABLE3": (T3_CAPTION, T3_TIER1, T3_TIER2, T3, T3_FOOT),
    "TABLE4": (T4_CAPTION, T4_TIER1, T4_TIER2, T4, T4_FOOT),
    "TABLE5": (T5_CAPTION, T5_TIER1, T5_TIER2, T5, T5_FOOT),
}


def build(sheet_name, caption, tier1, tier2, rows, footnotes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([caption])
    ws["A1"].font = Font(bold=True)
    ws.append(list(tier1))
    ws.append(list(tier2))
    for r in (2, 3):
        for c in ws[r]:
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical="bottom")
    for row in rows:
        ws.append(list(row))          # every cell a STRING, exactly as printed
    for f in footnotes:
        ws.append([f])
    widths = [34] + [15] * (len(tier2) - 1)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "B4"
    path = os.path.join(OUT, "Nudo_etal_1995_%s_snapshot.xlsx" % sheet_name)
    wb.save(path)
    return path, len(rows)


if __name__ == "__main__":
    for name, (cap, t1, t2, rows, foot) in TABLES.items():
        p, n = build(name, cap, t1, t2, rows, foot)
        assert all(len(r) == len(t2) for r in rows), "%s: ragged rows" % name
        print("%-8s %2d data rows  ->  %s" % (name, n, os.path.basename(p)))
