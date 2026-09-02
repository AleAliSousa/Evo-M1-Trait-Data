#!/usr/bin/env python3
"""Offline Python mirror of SensoryData_compiled.R (canonical).

CHECK FIXTURE, not a registry item (owner decision 2026-08-31): the Bath
"Sensory Data.xlsx" compilation is not a paper, so it is not built as a source
folder, has no registry row and no public TSV. Its long-format CSV serves as
the comparison fixture each Route-A primary-source folder is audited against
(match on the reference_short codes in the reference column).

The repo has no R in the Cowork sandbox; this mirror reproduces exactly what the
.R writes (R write.csv/write.table quoting: character fields quoted, numerics
bare, NA -> empty, floats via %.15g). If the .R is ever run, its output must be
byte-identical to this mirror's output; any diff is a bug in one of the two.

Input : Sensory Data.xlsx (frozen digital-native source, sheet "Data"), plus
        ../____Sensory_audiovisual/sensory_VA_offset_audit.csv (VA quarantine join)
Output: SensoryData_compiled.csv (long format, one row per species-row x trait value)
        ../__Public/comparative-data/COMPILATION%3Asensory_audiovisual_2022_Data.tsv
        SensoryData_compiled_definitions.csv
        SensoryData_compiled_references.csv
"""
import openpyxl, csv, os, re, sys

HERE = os.path.dirname(os.path.abspath(__file__))          # ____Sensory_audiovisual/SensoryData_compiled_check
SENS = os.path.dirname(HERE)                                # ____Sensory_audiovisual
ROOT = os.path.dirname(SENS)
SRC = os.path.join(SENS, "_incoming_Bath_archive_20260814", "Sensory Data.xlsx")
AUDIT = os.path.join(SENS, "sensory_VA_offset_audit.csv")
OUT_CSV = os.path.join(HERE, "SensoryData_compiled.csv")

BIRDS = {"Anas platyrhynchos", "Columba livia domestica",
         "Melopsittacus undulatus", "Taeniopygia guttata"}

# trait -> (value col, unit, reference col, comment col) ; method col 5 + substrate col 4 are VA-context
TRAITS = [
    ("visual_acuity_max",              3,  "c/deg",   7,  6),
    ("audible_freq_low_60dBSPL",       8,  "kHz",     9,  10),
    ("audible_freq_high_60dBSPL",      11, "kHz",     12, 13),
    ("hearing_range",                  14, "octaves", 15, None),
    ("interaural_distance_functional", 16, "us",      17, 18),
    ("best_sensitivity",               19, "dB",      20, 21),
    ("best_frequency",                 22, "kHz",     23, 24),
    ("sound_localization_threshold",   25, "deg",     26, 27),
    ("field_of_best_vision_width",     28, "deg",     29, 30),
    ("binocular_field",                31, "deg",     32, None),
    ("binaural_phase_cue_max_freq",    33, "kHz",     34, None),
    ("binaural_phase_cue",             35, "Y/N",     36, 37),
    ("binaural_intensity_difference_cue", 38, "Y/N",  39, 40),
    ("monaural_pinna_cues",            41, "Y/N",     42, None),
    ("trophic_level",                  43, "ordinal 1-5", 44, None),
]

def num_str(x):
    """R as.character() number formatting (15 significant digits)."""
    s = "%.15g" % x
    return s

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return num_str(v)
        return num_str(v)
    if isinstance(v, int):
        return str(v)
    return re.sub(r"\s+", " ", str(v)).strip()

def decimals_ge4(v):
    if not isinstance(v, float):
        return False
    s = "%.15g" % v
    return "." in s and len(s.split(".")[1]) >= 4

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Data"]

    # VA offset audit, keyed by source Excel row
    va_audit = {}
    with open(AUDIT, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            va_audit[int(row["SensoryData_excel_row"])] = row

    header = ["value_id", "Species_SensoryData", "common_name", "class",
              "source_excel_row", "trait", "value", "value_num", "unit",
              "method", "substrate", "reference", "comment",
              "value_origin", "qc_status", "qc_note"]
    out = []
    for r in range(2, 169):
        sp = ws.cell(r, 1).value
        if not sp:
            continue
        sp = str(sp).strip()
        common = cell_str(ws.cell(r, 2).value)
        klass = "Aves" if sp in BIRDS else "Mammalia"
        substrate = cell_str(ws.cell(r, 4).value)
        va_method = cell_str(ws.cell(r, 5).value)
        for trait, vc, unit, refc, comc in TRAITS:
            v = ws.cell(r, vc).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            ref = cell_str(ws.cell(r, refc).value)
            com = cell_str(ws.cell(r, comc).value) if comc else ""
            is_num = isinstance(v, (int, float))
            value = cell_str(v)
            value_num = num_str(float(v)) if is_num else ""
            method = va_method if trait == "visual_acuity_max" else ""
            origin = "digitised_from_figure" if decimals_ge4(v) else "published"
            qc, note = "ok", ""
            if trait == "visual_acuity_max":
                qc = "quarantined_va_offset"
                a = va_audit.get(r)
                if a:
                    note = "VA column quarantined pending re-derivation from Veilleux & Kirk 2014 / Kirk & Kay 2004; offset audit status: " + a["status"]
                    if a.get("SensoryData_value_is_Part1_value_of"):
                        note += "; value belongs upstream to: " + a["SensoryData_value_is_Part1_value_of"]
                else:
                    note = "VA column quarantined pending re-derivation; row absent from sensory_VA_offset_audit.csv"
            elif trait == "audible_freq_high_60dBSPL" and sp == "Delphinapterus leucas":
                qc, note = "curator_flag", "Notesissues #2: extrapolated Koay et al 1998 value conflicts with primary references"
            elif trait == "sound_localization_threshold" and sp == "Rattus norvegicus":
                qc, note = "curator_flag", "Notesissues #1: Heffner & Heffner 1992a value may be domestic vs albino Norway rat mix-up"
            value_id = "SensoryData_r%03d_%s" % (r, trait)
            out.append([value_id, sp, common, klass, str(r), trait, value,
                        value_num, unit, method, substrate, ref, com,
                        origin, qc, note])

    numeric_cols = {"value_num", "source_excel_row"}

    def write_r_style(path, sep):
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write(sep.join('"%s"' % h for h in header) + "\n")
            for row in out:
                fields = []
                for h, v in zip(header, row):
                    if h in numeric_cols:
                        fields.append(v)  # bare number or empty (na="")
                    else:
                        fields.append('"%s"' % v.replace('"', '""'))
                f.write(sep.join(fields) + "\n")

    write_r_style(OUT_CSV, ",")  # no public TSV: check fixture, not a registered item

    # reference_tables: definitions (10-col schema) + references (short code -> full citation)
    md = wb["Metadata"]
    defs_src = {}
    for rr in md.iter_rows(min_row=2, values_only=True):
        if rr[0]:
            defs_src[re.sub(r"\s+", " ", str(rr[0])).strip().lower()] = (
                re.sub(r"\s+", " ", str(rr[1] or "")).strip(),
                re.sub(r"\s+", " ", str(rr[2] or "")).strip())
    META_KEY = {  # trait -> Metadata sheet 'Variable' (lowercased)
        "visual_acuity_max": "maximum visual acuity",
        "audible_freq_low_60dBSPL": "lowest audible frequency at 60 db spl (khz)",
        "audible_freq_high_60dBSPL": "highest audible frequency at 60 db spl (khz)",
        "hearing_range": "hearing range (octaves)",
        "interaural_distance_functional": "functional interaural distance (µs)",
        "best_sensitivity": "best sensitivity (db)",
        "best_frequency": "best frequency (khz)",
        "sound_localization_threshold": "sound localization threshold (deg)",
        "field_of_best_vision_width": "width of field of best vision (deg)",
        "binocular_field": "binocular field (deg)",
        "binaural_phase_cue_max_freq": "highest frequency using binaural phase cue (khz)",
        "binaural_phase_cue": "binaural phase cue",
        "binaural_intensity_difference_cue": "binaural intensity-difference cue",
        "monaural_pinna_cues": "monoaural pinna cues",
        "trophic_level": "trophic level",
    }
    MEASURE = {"visual_acuity_max": "acuity.cdeg", "audible_freq_low_60dBSPL": "freq.kHz",
               "audible_freq_high_60dBSPL": "freq.kHz", "hearing_range": "range.octaves",
               "interaural_distance_functional": "time.us", "best_sensitivity": "level.dB",
               "best_frequency": "freq.kHz", "sound_localization_threshold": "angle.deg",
               "field_of_best_vision_width": "angle.deg", "binocular_field": "angle.deg",
               "binaural_phase_cue_max_freq": "freq.kHz", "binaural_phase_cue": "binary",
               "binaural_intensity_difference_cue": "binary", "monaural_pinna_cues": "binary",
               "trophic_level": "ordinal"}
    dpath = os.path.join(HERE, "SensoryData_compiled_definitions.csv")
    with open(dpath, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(["Code", "Definition", "Structure", "Measure", "Stat", "role",
                    "taxon", "Reference", "Note", "Source Note"])
        for trait, *_ in TRAITS:
            d, src = defs_src.get(META_KEY[trait], ("", ""))
            note = ""
            if trait == "visual_acuity_max":
                note = "QUARANTINED: 24-row offset error, see ReadMe and sensory_VA_offset_audit.csv"
            w.writerow([trait, d, "", MEASURE[trait], "value", "secondary",
                        "Mammalia (4 Aves rows registered, not compiled)",
                        "SensoryData_compiled", note, src])
        for code, defi in [("Species_SensoryData", "Species name verbatim as in Sensory Data.xlsx"),
                           ("method", "Visual acuity method: B = behavioral, A = anatomical"),
                           ("substrate", "Measurement medium, air or water"),
                           ("value_origin", "published, or digitised_from_figure (>=4 decimals; chiefly Koay et al 1998 Fig. 6, Heffner 2018)"),
                           ("qc_status", "ok / quarantined_va_offset / curator_flag"),
                           ("Method:units", "Values kept in source units (performance/psychophysics measure class; no volume/mass conversions apply)")]:
            w.writerow([code, defi, "", "", "", "info", "", "SensoryData_compiled", "", ""])

    rs = wb["Reference"]
    rpath = os.path.join(HERE, "SensoryData_compiled_references.csv")
    with open(rpath, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(["reference_short", "full_citation", "note"])
        for rr in rs.iter_rows(min_row=2, values_only=True):
            if rr[0] and rr[1]:
                w.writerow([re.sub(r"\s+", " ", str(rr[0])).strip(),
                            re.sub(r"\s+", " ", str(rr[1])).strip(),
                            re.sub(r"\s+", " ", str(rr[10] or rr[2] or "")).strip()])

    print("rows written:", len(out))
    from collections import Counter
    print(Counter(r[5] for r in out))
    print(Counter(r[14] for r in out))

if __name__ == "__main__":
    main()
