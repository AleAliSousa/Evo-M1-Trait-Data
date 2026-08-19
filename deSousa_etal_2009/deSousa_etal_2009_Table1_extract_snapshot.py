#!/usr/bin/env python3
"""Build the de Sousa et al. 2009 Table 1 snapshot (Cereb Cortex 20(4):966-981, Advance Access p. 2).

Faithful capture per __HOWTO_make_a_snapshot.md: the printed caption, the three-tier
column header, footnote markers kept attached to the strings that carry them
(``Homo sapiensa,b``, ``EQe``, ``area (mm2)f``), printed blanks left blank, printed
row order and original units (kg / g / cm3 / mm3 / mm2) unchanged. Values are stored
as the printed *strings* so trailing zeros ("1.20", "16.0") and the literal "NA"
survive verbatim; the reformat (deSousa_etal_2009_Table1.R) does the typing.

Nothing is corrected here. Two things that look wrong are carried as printed and
flagged in the README instead:
  * ``ppz`` (Zahlia) neocortex 279 cm3 is the value de Sousa et al. 2010 Table 1
    prints for a *different* bonobo (YN86-137); Zahlia's own is 214.4.
  * brain mass is rounded relative to 2010 Table 1 (58 vs 57.6; 360 vs 359.5).

Run:  python3 deSousa_etal_2009_Table1_extract_snapshot.py [--verify]
--verify re-reads the PDF with pdfplumber and asserts every transcribed token is
actually on the page in the expected row (it is the transcription audit; it needs
pdfplumber and the PDF, and is skipped silently if either is missing).
"""
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font

OUT = os.path.dirname(os.path.abspath(__file__))
PDF = os.path.join(OUT, "desousa_etal_2009_AdvanceAccess.pdf")
XLSX = os.path.join(OUT, "deSousa_etal_2009_Table1_snapshot.xlsx")

CAPTION = "Table 1"
SUBCAPTION = "Samples used in analyses of V1, V2, and VP"

# Three printed header tiers. Superscript footnote letters are kept attached
# (EQe, area (mm2)f) exactly as they read on the page.
TIER1 = ["Species", "code", "archive", "sex", "age (yrs)", "body mass", "brain mass",
         "EQe", "neocortex", "left V1 vol", "left LGN", "optic nerve", "eye half surface"]
TIER2 = ["", "", "number", "", "", "(kg)", "(g)",
         "", "volume", "(mm3)", "vol (mm3)", "cross sectional", "area (mm2)f"]
TIER3 = ["", "", "", "", "", "", "",
         "", "(cm3)", "", "", "area (mm2)f", ""]

# Printed row order. "" = printed blank.
ROWS = [
    ["Gorilla gorilla",     "ggy", "YN82-140",  "F",  "20",   "85", "376",  "1.20", "254", "4044", "150", "17.6", "1774"],
    ["Hylobates lar",       "hld", "Disco",     "F",  "22",   "7",  "120",  "2.54", "77",  "2292", "90",  "12.7", "1299"],
    ["Homo sapiensa,b",     "hs5", "54491",     "F",  "79",   "63", "1350", "5.41", "974", "7587", "186", "22.8", "1855"],
    ["Homo sapiensa,b",     "hs6", "6895",      "F",  "79",   "63", "1110", "4.45", "974", "7013", "156", "22.8", "1855"],
    ["Macaca fascicularis", "mf2", "ma22",      "M",  "3",    "3",  "58",   "2.31", "",    "1357", "46",  "8.4",  "985"],
    ["Pongo pygmaeus",      "ouy", "YN85-38",   "M",  "16.5", "58", "369",  "1.56", "269", "3504", "92",  "16.1", "1282"],
    ["Pan paniscusc",       "ppz", "Zahlia",    "F",  "11",   "33", "324",  "2.09", "279", "5687", "130", "",     ""],
    ["Pan troglodytes",     "ptb", "Bathsheba", "F",  "24",   "80", "360",  "1.20", "263", "4705", "168", "16.0", "1446"],
    ["Pan troglodytesd",    "ptd", "1548",      "NA", "NA",   "51", "387",  "1.82", "198", "2799", "86",  "16.0", "1446"],
]

# Printed footnotes. "n = 8" is printed with the font's "5" glyph for "=" ("n 5 8");
# it is transcribed as the "=" it stands for, which is the only glyph normalisation here.
FOOTNOTES = [
    "a Used same sex species mean value for body weight (Zilles 1972).",
    "b Used combined sex mean human neocortex value (n = 8) based on unpublished data provided by Carol MacLeod.",
    "c Used same sex species mean value for body weight (Jungers and Susman 1984).",
    "d Used combined sex species mean values for brain and body weight (Herndon et al. 1999.).",
    "e Encephalization quotient (EQ) after Martin (1981) and Ruff et al. (1997).",
    "f Species mean data from Stephan and Frahm 1981.",
]


def write_snapshot():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table1"

    ws.cell(row=1, column=1, value=CAPTION).font = Font(bold=True)
    ws.cell(row=2, column=1, value=SUBCAPTION)
    for i, tier in enumerate((TIER1, TIER2, TIER3), start=3):
        for j, v in enumerate(tier, start=1):
            if v:
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(bold=True)
                c.alignment = Alignment(wrap_text=True, vertical="bottom")
    for i, row in enumerate(ROWS, start=6):
        for j, v in enumerate(row, start=1):
            if v != "":
                ws.cell(row=i, column=j, value=v)
    for i, note in enumerate(FOOTNOTES, start=16):
        ws.cell(row=i, column=1, value=note).font = Font(size=9)

    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 10
    for col in "FGHIJKLM":
        ws.column_dimensions[col].width = 13
    ws.freeze_panes = "A6"
    wb.save(XLSX)
    print("wrote", XLSX, "-", len(ROWS), "data rows,", len(TIER1), "columns")


def verify_against_pdf():
    """Assert every transcribed token appears in the matching printed row of the PDF."""
    try:
        import pdfplumber
    except ImportError:
        print("verify skipped: pdfplumber not installed")
        return
    if not os.path.exists(PDF):
        print("verify skipped: PDF not found")
        return
    with pdfplumber.open(PDF) as pdf:
        words = pdf.pages[1].extract_words()
    # Cluster words into printed rows by baseline. Superscript footnote letters sit ~1.2 pt
    # above their own row's baseline, so an exact-top grouping would split them off.
    body, current, top0 = [], [], None
    for w in sorted((w for w in words if 110 < w["top"] < 190), key=lambda w: (w["top"], w["x0"])):
        if top0 is None or w["top"] - top0 > 3:
            if current:
                body.append(" ".join(t for _, t in sorted(current)))
            current, top0 = [], w["top"]
        current.append((w["x0"], w["text"]))
    if current:
        body.append(" ".join(t for _, t in sorted(current)))
    bad = []
    for row in ROWS:
        printed = next((b for b in body if row[1] in b.split()), None)
        if printed is None:
            bad.append((row[1], "row not found on page"))
            continue
        # the PDF drops spaces inside words and sets footnote letters as separate words
        flat = printed.replace(" ", "")
        for v in row:
            if v and v.replace(" ", "") not in flat:
                bad.append((row[1], v))
    print("verify: %d/%d rows matched the PDF" % (len(ROWS) - len({b[0] for b in bad}), len(ROWS)))
    if bad:
        raise SystemExit("transcription mismatch: %r" % bad)


if __name__ == "__main__":
    write_snapshot()
    if "--verify" in sys.argv:
        verify_against_pdf()
